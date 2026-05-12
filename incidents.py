from flask import Flask, render_template_string, request, redirect, session, send_file, abort
import csv
from datetime import date, datetime
import os
import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from werkzeug.utils import secure_filename
from PIL import Image as PILImage

app = Flask(__name__)
app.secret_key = "supersecretkey"  # Needed for session management

CSV_FILE = "incidents.csv"
UPLOAD_FOLDER = "uploads"
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER

VIEW_PASSWORD = "mypassword"  # Only you can see the incidents
ALLOWED_EXTENSIONS = {"png", "jpg", "jpeg", "gif"}
CSV_HEADERS = ["ID","Title","Description","Severity","Status","DateCreated","Assignee","Screenshots","Resolution","ClosedDate"]

# Ensure CSV file exists with header
if not os.path.exists(CSV_FILE):
    with open(CSV_FILE, "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(CSV_HEADERS)

# Load incidents from CSV
def load_incidents():
    incidents = []
    with open(CSV_FILE, "r", newline="") as f:
        reader = csv.reader(f)
        next(reader)  # Skip header
        for row in reader:
            while len(row) < len(CSV_HEADERS):
                row.append("")
            incidents.append(row)
    return incidents

# Save all incidents to CSV
def save_all_incidents(incidents):
    with open(CSV_FILE, "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(CSV_HEADERS)
        writer.writerows(incidents)

# Save a new incident to CSV
def save_incident(title, description, severity, assignee, files):
    incidents = load_incidents()
    next_id = len(incidents) + 1
    filenames = []
    if files:
        for file in files:
            if file and "." in file.filename and file.filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS:
                filename = f"{next_id}_{secure_filename(file.filename)}"
                filepath = os.path.join(app.config["UPLOAD_FOLDER"], filename)
                file.save(filepath)
                filenames.append(filename)
    today_str = date.today().strftime("%Y-%m-%d")
    with open(CSV_FILE, "a", newline="") as f:
        writer = csv.writer(f)
        writer.writerow([next_id, title, description, severity, "Open", today_str, assignee, "|".join(filenames), "", ""])

# Update an incident's status
def update_incident_status(incident_id, new_status, resolution="", closed_date=""):
    incidents = load_incidents()
    for i, inc in enumerate(incidents):
        if inc[0] == str(incident_id):
            incidents[i][4] = new_status
            incidents[i][8] = resolution
            incidents[i][9] = closed_date if new_status == "Closed" else ""
            break
    save_all_incidents(incidents)

# Delete an incident
def delete_incident(incident_id):
    incidents = load_incidents()
    incidents = [inc for inc in incidents if inc[0] != str(incident_id)]
    save_all_incidents(incidents)

# Excel export function with per-incident image sheets and DD-MM-YYYY date formatting
def export_incidents_excel():
    incidents = load_incidents()
    wb = Workbook()
    ws = wb.active
    ws.title = "Incidents"

    headers = ["ID", "Title", "Description", "Severity", "Status", "DateCreated", "Name", "Images Sheet", "Resolution", "Closed Date"]
    ws.append(headers)

    for col in range(1, len(headers)+1):
        ws.cell(row=1, column=col).font = Font(bold=True)
        ws.cell(row=1, column=col).alignment = Alignment(horizontal="center")

    for i in incidents:
        images_sheet = f"Incident {i[0]}" if i[7] else ""
        try:
            date_obj = datetime.strptime(i[5], "%Y-%m-%d")
        except:
            date_obj = i[5]
        try:
            closed_date_obj = datetime.strptime(i[9], "%Y-%m-%d") if i[9] else ""
        except:
            closed_date_obj = i[9]
        ws.append([i[0], i[1], i[2], i[3], i[4], date_obj, i[6], images_sheet, i[8], closed_date_obj])
        if isinstance(date_obj, datetime):
            ws.cell(row=ws.max_row, column=6).number_format = "DD-MM-YYYY"
        if isinstance(closed_date_obj, datetime):
            ws.cell(row=ws.max_row, column=10).number_format = "DD-MM-YYYY"

    severity_colors = {"Low": "D6D8D9", "Medium": "CCE5FF", "High": "FFF3CD", "Critical": "F8D7DA"}
    for row in range(2, ws.max_row+1):
        sev = ws.cell(row=row, column=4).value
        sev_color = severity_colors.get(sev, "FFFFFF")
        for col in range(1, len(headers)+1):
            ws.cell(row=row, column=col).fill = PatternFill(start_color=sev_color, end_color=sev_color, fill_type="solid")
            ws.cell(row=row, column=col).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    column_widths = [5, 25, 40, 10, 15, 12, 20, 20, 40, 12]
    for idx, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    for i in incidents:
        if not i[7]:
            continue
        sheet_name = f"Incident {i[0]}"[:31]
        ws_img = wb.create_sheet(title=sheet_name)
        ws_img["A1"] = f"Incident {i[0]} - {i[1]}"
        ws_img["A1"].font = Font(bold=True)
        row_cursor = 3
        for img_name in i[7].split('|'):
            path = os.path.join(app.config["UPLOAD_FOLDER"], img_name)
            if os.path.exists(path):
                try:
                    img = XLImage(path)
                    with PILImage.open(path) as pil_img:
                        max_w, max_h = 300, 200
                        orig_w, orig_h = pil_img.size
                        ratio = min(max_w / orig_w, max_h / orig_h)
                        img.width = int(orig_w * ratio)
                        img.height = int(orig_h * ratio)
                    ws_img.add_image(img, f"A{row_cursor}")
                    ws_img.row_dimensions[row_cursor].height = img.height * 0.75
                    row_cursor += 10
                except:
                    ws_img[f"A{row_cursor}"] = f"Error loading: {img_name}"
                    row_cursor += 2

    for row in range(2, ws.max_row+1):
        sheet_name = ws.cell(row=row, column=8).value
        if sheet_name:
            ws.cell(row=row, column=8).hyperlink = f"#{sheet_name}!A1"
            ws.cell(row=row, column=8).font = Font(color="0000FF", underline="single")

    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    return file_stream

@app.route("/download_excel")
def download_excel():
    file_stream = export_incidents_excel()
    return send_file(
        file_stream,
        as_attachment=True,
        download_name="incidents.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.route("/uploads/<path:filename>")
def uploaded_file(filename):
    safe_filename = secure_filename(os.path.basename(filename))

    file_path = os.path.realpath(
        os.path.join(app.config["UPLOAD_FOLDER"], safe_filename)
    )

    upload_folder = os.path.realpath(app.config["UPLOAD_FOLDER"])

    if os.path.commonpath([upload_folder, file_path]) != upload_folder:
        abort(403)

    if not os.path.isfile(file_path):
        abort(404)

    return send_file(file_path)

@app.route("/", methods=["GET", "POST"])
def index():
    show_incidents = session.get("can_view", False)

    if request.method == "POST" and "view_password" in request.form:
        if request.form["view_password"] == VIEW_PASSWORD:
            session["can_view"] = True
        return redirect("/")

    if request.method == "POST" and "title" in request.form:
        save_incident(
            request.form["title"],
            request.form["description"],
            request.form["severity"],
            request.form["assignee"],
            request.files.getlist("screenshots")
        )
        return redirect("/")

    incidents = list(reversed(load_incidents())) if show_incidents else []

    return render_template_string("""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<title>Company Incident Tracker</title>
<link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/css/bootstrap.min.css" rel="stylesheet">
<style>
body { background-color: #b6e3ff; }
.container { margin-top: 40px; }
.row-low { background-color: #b6e3ff; }
.row-medium { background-color: #e7f1ff; }
.row-high { background-color: #fff3cd; color: #856404; }
.row-critical { background-color: #f8d7da; color: #721c24; }
.status-open { background-color: rgba(25, 135, 84, 0.15); }
.status-inprogress { background-color: rgba(13, 202, 240, 0.15); }
.status-closed { background-color: rgba(108, 117, 125, 0.15); }
.badge-severity-low { background-color: #6c757d; }
.badge-severity-medium { background-color: #0d6efd; }
.badge-severity-high { background-color: #ffc107; color: #212529; }
.badge-severity-critical { background-color: #dc3545; }
.badge-status-open { background-color: #198754; }
.badge-status-inprogress { background-color: #0dcaf0; color: #212529; }
.badge-status-closed { background-color: #6c757d; }
.easter-egg-button {
    position: fixed;
    bottom: 10px;
    right: 10px;
    background-color: #0d6efd;
    color: white;
    border: none;
    border-radius: 20px;
    padding: 4px 10px;
    font-size: 10px;
    opacity: 0.7;
    cursor: pointer;
    z-index: 9999;
}
.easter-egg-button:hover {
    opacity: 1;
}
</style>
</head>
<body>
<div class="container">
<h2 class="mb-4 text-center">Company Incident Tracker</h2>
<!-- Add Incident Form -->
<div class="card mb-4 shadow-sm">
<div class="card-body">
<form method="post" enctype="multipart/form-data">
<div class="mb-3"><input name="assignee" placeholder="Name" class="form-control"></div>
<div class="mb-3"><input name="title" placeholder="Title of incident" class="form-control" required></div>
<div class="mb-3"><textarea name="description" placeholder="Description" class="form-control" rows="3"></textarea></div>
<div class="mb-3">
<select name="severity" class="form-select">
<option>Low</option><option>Medium</option><option>High</option><option>Critical</option>
</select>
</div>
<div class="mb-3"><input type="file" name="screenshots" multiple class="form-control"></div>
<button type="submit" class="btn btn-primary">Add Incident</button>
</form>
</div>
</div>
{% if not session.get('can_view') %}
<div class="card mb-4 shadow-sm">
<div class="card-body">
<form method="post">
<div class="mb-3">
<input name="view_password" type="password" placeholder="Enter password to view incidents" class="form-control">
</div>
<button type="submit" class="btn btn-secondary">View Incidents</button>
</form>
</div>
</div>
{% else %}
<div class="d-flex justify-content-between align-items-center mb-2">
<h3>Incidents</h3>
<div>
<a href="/download_excel" class="btn btn-success btn-sm me-2">Download Excel</a>
<form method="post" action="/logout" style="display:inline">
<button class="btn btn-danger btn-sm">Logout</button>
</form>
</div>
</div>
<div class="table-responsive">
<table class="table table-hover table-striped align-middle shadow-sm bg-white">
<thead class="table-light">
<tr>
<th>ID</th><th>Title</th><th>Severity</th><th>Status</th><th>Date</th><th>Assignee</th><th>Resolution</th><th>Closed Date</th><th>Actions</th>
</tr>
</thead>
<tbody>
{% for i in incidents %}
{% set row_class = '' %}
{% if i[3] == 'Low' %} {% set row_class = 'row-low' %}
{% elif i[3] == 'Medium' %} {% set row_class = 'row-medium' %}
{% elif i[3] == 'High' %} {% set row_class = 'row-high' %}
{% elif i[3] == 'Critical' %} {% set row_class = 'row-critical' %} {% endif %}
{% set status_class = '' %}
{% if i[4] == 'Open' %} {% set status_class = 'status-open' %}
{% elif i[4] == 'In Progress' %} {% set status_class = 'status-inprogress' %}
{% elif i[4] == 'Closed' %} {% set status_class = 'status-closed' %} {% endif %}
<tr class="{{ row_class }} {{ status_class }}">
<td>{{ i[0] }}</td>
<td>{{ i[1] }}</td>
<td>
{% if i[3] == 'Low' %}<span class="badge badge-severity-low">Low</span>
{% elif i[3] == 'Medium' %}<span class="badge badge-severity-medium">Medium</span>
{% elif i[3] == 'High' %}<span class="badge badge-severity-high">High</span>
{% elif i[3] == 'Critical' %}<span class="badge badge-severity-critical">Critical</span>
{% endif %}
</td>
<td>
{% if i[4] == 'Open' %}<span class="badge badge-status-open">Open</span>
{% elif i[4] == 'In Progress' %}<span class="badge badge-status-inprogress">In Progress</span>
{% elif i[4] == 'Closed' %}<span class="badge badge-status-closed">Closed</span>
{% else %}<span class="badge bg-secondary">{{ i[4] }}</span>{% endif %}
</td>
<td>{{ i[5] }}</td>
<td>{{ i[6] }}</td>
<td>{{ i[8] }}</td>
<td>{{ i[9] }}</td>
<td>
{% if i[4] != 'Closed' %}
<form method="post" action="/close_incident/{{ i[0] }}" style="display:inline">
<input type="text" name="resolution" placeholder="How was it fixed?" class="form-control form-control-sm mb-1" required>
<input type="date" name="closed_date" class="form-control form-control-sm mb-1" required>
<button class="btn btn-sm btn-warning">Close</button>
</form>
{% endif %}
<form method="post" action="/delete_incident/{{ i[0] }}" style="display:inline" onsubmit="return confirm('Are you sure you want to delete this incident?');"><button class="btn btn-sm btn-danger">Delete</button></form>
</td>
</tr>
{% endfor %}
</tbody>
</table>
</div>
{% endif %}
</div>

<div style="
    position: fixed;
    bottom: 10px;
    left: 10px;
    font-size: 11px;
    color: #6c757d;
    opacity: 0.8;
    z-index: 9999;
">
    © 2026 Gerben Rohof — All rights reserved
</div>

<button class="easter-egg-button" onclick="showEasterEgg()">
    Click here!
</button>

<div id="easterEggModal" style="
    display:none;
    position:fixed;
    top:0;
    left:0;
    width:100%;
    height:100%;
    background:rgba(0,0,0,0.7);
    justify-content:center;
    align-items:center;
    z-index:10000;
">
    <div style="
        background:white;
        padding:20px;
        border-radius:15px;
        text-align:center;
        max-width:500px;
        box-shadow:0 0 20px rgba(0,0,0,0.3);
    ">
        <h3>Geppie was here 😎</h3>

        <img src="/static/geppie.gif"
             style="
                max-width:100%;
                border-radius:10px;
                margin-top:10px;
             ">

        <br><br>

        <button onclick="closeEasterEgg()" class="btn btn-primary">
            Close
        </button>
    </div>
</div>

<script>
function showEasterEgg() {
    document.getElementById("easterEggModal").style.display = "flex";
}

function closeEasterEgg() {
    document.getElementById("easterEggModal").style.display = "none";
}
</script>
</body>
</html>""", incidents=incidents)

@app.route("/close_incident/<int:incident_id>", methods=["POST"])
def close_incident(incident_id):
    resolution = request.form.get("resolution", "")
    closed_date = request.form.get("closed_date", date.today().strftime("%Y-%m-%d"))
    update_incident_status(incident_id, "Closed", resolution, closed_date)
    return redirect("/")

@app.route("/delete_incident/<int:incident_id>", methods=["POST"])
def remove_incident(incident_id):
    delete_incident(incident_id)
    return redirect("/")

@app.route("/logout", methods=["POST"])
def logout():
    session.pop("can_view", None)
    return redirect("/")

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=42069)
