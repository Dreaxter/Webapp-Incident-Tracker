from flask import Flask, render_template_string, request, redirect, session, send_file
import csv
from datetime import date
import os
import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment

app = Flask(__name__)
app.secret_key = "supersecretkey"  # Needed for session management

CSV_FILE = "incidents.csv"
VIEW_PASSWORD = "mypassword"  # Only you can see the incidents

# Ensure CSV file exists with header
if not os.path.exists(CSV_FILE):
    with open(CSV_FILE, "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["ID","Title","Description","Severity","Status","DateCreated","Assignee"])

# Load incidents from CSV
def load_incidents():
    incidents = []
    with open(CSV_FILE, "r", newline="") as f:
        reader = csv.reader(f)
        next(reader)  # Skip header
        for row in reader:
            incidents.append(row)
    return incidents

# Save a new incident to CSV
def save_incident(title, description, severity, assignee):
    incidents = load_incidents()
    next_id = len(incidents) + 1
    with open(CSV_FILE, "a", newline="") as f:
        writer = csv.writer(f)
        writer.writerow([next_id, title, description, severity, "Open", date.today().isoformat(), assignee])

# Update an incident's status
def update_incident_status(incident_id, new_status):
    incidents = load_incidents()
    for i, inc in enumerate(incidents):
        if inc[0] == str(incident_id):
            incidents[i][4] = new_status
            break
    with open(CSV_FILE, "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["ID","Title","Description","Severity","Status","DateCreated","Assignee"])
        writer.writerows(incidents)

# Delete an incident
def delete_incident(incident_id):
    incidents = load_incidents()
    incidents = [inc for inc in incidents if inc[0] != str(incident_id)]
    with open(CSV_FILE, "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["ID","Title","Description","Severity","Status","DateCreated","Assignee"])
        writer.writerows(incidents)

# Excel export function
def export_incidents_excel():
    incidents = load_incidents()
    wb = Workbook()
    ws = wb.active
    ws.title = "Incidents"

    headers = ["ID", "Title", "Description", "Severity", "Status", "DateCreated", "Assignee"]
    ws.append(headers)

    # Style header
    for col in range(1, len(headers)+1):
        ws.cell(row=1, column=col).font = Font(bold=True)
        ws.cell(row=1, column=col).alignment = Alignment(horizontal="center")

    severity_colors = {"Low": "D6D8D9", "Medium": "CCE5FF", "High": "FFF3CD", "Critical": "F8D7DA"}
    status_colors = {"Open": ("198754", "FFFFFF"), "In Progress": ("0DCaf0", "000000"), "Closed": ("6C757D", "FFFFFF")}

    for i in incidents:
        ws.append(i)
        row = ws.max_row

        sev_color = severity_colors.get(i[3], "FFFFFF")
        for col in range(1, 8):
            ws.cell(row=row, column=col).fill = PatternFill(start_color=sev_color, end_color=sev_color, fill_type="solid")
            ws.cell(row=row, column=col).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        status = i[4]
        if status in status_colors:
            bg, font_color = status_colors[status]
            cell = ws.cell(row=row, column=5)
            cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
            cell.font = Font(color=font_color, bold=True)

        if i[3] == "Critical" and i[4] == "Open":
            for col in range(1, 8):
                ws.cell(row=row, column=col).font = Font(bold=True, color="9C0006")

    column_widths = [5, 25, 40, 10, 15, 12, 20]
    for idx, width in enumerate(column_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = width

    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    return file_stream

@app.route("/download_excel")
def download_excel():
    file_stream = export_incidents_excel()
    return send_file(
        file_stream,
        as_attachment=True,
        download_name="incidents.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.route("/", methods=["GET", "POST"])
def index():
    show_incidents = session.get("can_view", False)

    # Password to view incidents
    if request.method == "POST" and "view_password" in request.form:
        if request.form["view_password"] == VIEW_PASSWORD:
            session["can_view"] = True
        return redirect("/")

    # Add new incident
    if request.method == "POST" and "title" in request.form:
        save_incident(
            request.form["title"],
            request.form["description"],
            request.form["severity"],
            request.form["assignee"]
        )
        return redirect("/")

    incidents = []
    if show_incidents:
        incidents = list(reversed(load_incidents()))

    return render_template_string("""
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <title>Company Incident Tracker</title>
    <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/css/bootstrap.min.css" rel="stylesheet">
    <style>
        body { background-color: #f8f9fa; }
        .container { margin-top: 40px; }
        .row-low { background-color: #f8f9fa; }
        .row-medium { background-color: #e7f1ff; }
        .row-high { background-color: #fff3cd; color: #856404; }
        .row-critical { background-color: #f8d7da; color: #721c24; }
        .status-open { background-color: rgba(25, 135, 84, 0.15); }
        .status-inprogress { background-color: rgba(13, 202, 240, 0.15); }
        .status-closed { background-color: rgba(108, 117, 125, 0.15); }
        .badge-severity-low { background-color: #6c757d; }
        .badge-severity-medium { background-color: #0d6efd; }
        .badge-severity-high { background-color: #ffc107; color: #212529; }
        .badge-severity-critical { background-color: #dc3545; }
        .badge-status-open { background-color: #198754; }
        .badge-status-inprogress { background-color: #0dcaf0; color: #212529; }
        .badge-status-closed { background-color: #6c757d; }
    </style>
</head>
<body>
<div class="container">
    <h2 class="mb-4 text-center">Company Incident Tracker</h2>

    <!-- Add Incident Form -->
    <div class="card mb-4 shadow-sm">
        <div class="card-body">
            <form method="post">
                <div class="mb-3"><input name="title" placeholder="Title" class="form-control" required></div>
                <div class="mb-3"><textarea name="description" placeholder="Description" class="form-control" rows="3"></textarea></div>
                <div class="mb-3">
                    <select name="severity" class="form-select">
                        <option>Low</option><option>Medium</option><option>High</option><option>Critical</option>
                    </select>
                </div>
                <div class="mb-3"><input name="assignee" placeholder="Assignee" class="form-control"></div>
                <button type="submit" class="btn btn-primary">Add Incident</button>
            </form>
        </div>
    </div>

    {% if not session.get('can_view') %}
        <div class="card mb-4 shadow-sm">
            <div class="card-body">
                <form method="post">
                    <div class="mb-3">
                        <input name="view_password" type="password" placeholder="Enter password to view incidents" class="form-control">
                    </div>
                    <button type="submit" class="btn btn-secondary">View Incidents</button>
                </form>
            </div>
        </div>
    {% else %}
        <div class="d-flex justify-content-between align-items-center mb-2">
            <h3>Incidents</h3>
            <div>
                <a href="/download_excel" class="btn btn-success btn-sm me-2">Download Excel</a>
                <form method="post" action="/logout" style="display:inline">
                    <button class="btn btn-danger btn-sm">Logout</button>
                </form>
            </div>
        </div>

        <div class="table-responsive">
            <table class="table table-hover table-striped align-middle shadow-sm bg-white">
                <thead class="table-light">
                    <tr>
                        <th>ID</th><th>Title</th><th>Severity</th><th>Status</th><th>Date</th><th>Assignee</th><th>Actions</th>
                    </tr>
                </thead>
                <tbody>
                    {% for i in incidents %}
                        {% set row_class = '' %}
                        {% if i[3] == 'Low' %} {% set row_class = 'row-low' %}
                        {% elif i[3] == 'Medium' %} {% set row_class = 'row-medium' %}
                        {% elif i[3] == 'High' %} {% set row_class = 'row-high' %}
                        {% elif i[3] == 'Critical' %} {% set row_class = 'row-critical' %} {% endif %}

                        {% set status_class = '' %}
                        {% if i[4] == 'Open' %} {% set status_class = 'status-open' %}
                        {% elif i[4] == 'In Progress' %} {% set status_class = 'status-inprogress' %}
                        {% elif i[4] == 'Closed' %} {% set status_class = 'status-closed' %} {% endif %}

                        <tr class="{{ row_class }} {{ status_class }}">
                            <td>{{ i[0] }}</td>
                            <td>{{ i[1] }}</td>
                            <td>
                                {% if i[3] == 'Low' %}
                                    <span class="badge badge-severity-low">Low</span>
                                {% elif i[3] == 'Medium' %}
                                    <span class="badge badge-severity-medium">Medium</span>
                                {% elif i[3] == 'High' %}
                                    <span class="badge badge-severity-high">High</span>
                                {% elif i[3] == 'Critical' %}
                                    <span class="badge badge-severity-critical">Critical</span>
                                {% endif %}
                            </td>
                            <td>
                                {% if i[4] == 'Open' %}
                                    <span class="badge badge-status-open">Open</span>
                                {% elif i[4] == 'In Progress' %}
                                    <span class="badge badge-status-inprogress">In Progress</span>
                                {% elif i[4] == 'Closed' %}
                                    <span class="badge badge-status-closed">Closed</span>
                                {% else %}
                                    <span class="badge bg-secondary">{{ i[4] }}</span>
                                {% endif %}
                            </td>
                            <td>{{ i[5] }}</td>
                            <td>{{ i[6] }}</td>
                            <td>
                                {% if i[4] != 'Closed' %}
                                    <form method="post" action="/close_incident/{{ i[0] }}" style="display:inline">
                                        <button class="btn btn-sm btn-warning">Close</button>
                                    </form>
                                {% endif %}
                                <form method="post" action="/delete_incident/{{ i[0] }}" style="display:inline" onsubmit="return confirm('Are you sure you want to delete this incident?');">
                                    <button class="btn btn-sm btn-danger">Delete</button>
                                </form>
                            </td>
                        </tr>
                    {% endfor %}
                </tbody>
            </table>
        </div>
    {% endif %}
</div>
</body>
</html>
""", incidents=incidents)

@app.route("/close_incident/<int:incident_id>", methods=["POST"])
def close_incident(incident_id):
    update_incident_status(incident_id, "Closed")
    return redirect("/")

@app.route("/delete_incident/<int:incident_id>", methods=["POST"])
def remove_incident(incident_id):
    delete_incident(incident_id)
    return redirect("/")

@app.route("/logout", methods=["POST"])
def logout():
    session.pop("can_view", None)
    return redirect("/")

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=42069)
